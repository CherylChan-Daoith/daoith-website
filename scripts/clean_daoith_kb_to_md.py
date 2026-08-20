#!/usr/bin/env python3
"""Clean 道一跨境电商知识库 source files into Dify-friendly Markdown.

Sources (Desktop):
  - 跨境外贸财税300问final.docx  → 按篇拆分 Q&A
  - 审核规则-佛山中山.xlsx       → 填平合并单元格后的规则表
  - 首次申报出口退税资料清单5.0.xlsx → 资料清单+填报说明
  - 备案单证规则-20240806.xmind  → 按脑图 sheet 拆成独立规则篇

Dify notes:
  - 不用超大合集；按主题分文件
  - 问题用 ## 标题，便于切片
  - 去掉营销 CTA；保留规则正文与表格
"""
from __future__ import annotations

import re
import shutil
import zipfile
from collections import defaultdict
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
from docx import Document
from docx.document import Document as DocumentType
from docx.oxml.ns import qn
from docx.table import Table
from docx.text.paragraph import Paragraph

ROOT = Path(__file__).resolve().parents[1]
SRC = Path.home() / "Desktop" / "DIFY知识库" / "道一跨境电商知识库"
OUT = ROOT / "data" / "daoith-kb-md"
DESKTOP_MD = SRC / "md"
AS_OF = "2026-08"

SCOPE = (
    "> **知识库用途**：本文件由道一内部资料清洗为 Markdown，供检索问答。"
    "规则可能更新，请以现行法规与内部最新制度为准。本文不构成对外法律意见。\n"
)

CTA_PATTERNS = [
    re.compile(r"如需就具体或个别情形进行单独咨询"),
    re.compile(r"点击这里预约"),
    re.compile(r"贸E税专家"),
    re.compile(r"一对一沟通"),
]

PARTS = [
    (1, 80, "01-财税合规篇", "第一部分：财税合规篇（第1～80问）"),
    (81, 160, "02-出口退税篇", "第二部分：出口退税篇（第81～160问）"),
    (161, 210, "03-财税规划篇", "第三部分：财税规划篇（第161～210问）"),
    (211, 250, "04-账务处理篇", "第四部分：账务处理篇（第211～250问）"),
    (251, 300, "05-外汇关务篇", "第五部分：外汇、关务篇（第251～300问）"),
]


def clean_text(s: str) -> str:
    s = (s or "").replace("\r\n", "\n").replace("\r", "\n")
    s = s.replace("\u00a0", " ").strip()
    return s


def rewrite_merchant_terms(text: str) -> str:
    """对外口径：用「地区商家」替代内部「子公司」说法。"""
    if not text:
        return text
    # 先处理固定短语（长词优先）
    reps = [
        ("其它子公司", "其他地区商家"),
        ("其他子公司", "其他地区商家"),
        ("所有子公司", "所有地区商家"),
        ("广州和中山子公司", "广州和中山地区商家"),
        ("泉州/福建子公司", "泉州/福建地区商家"),
        ("华东浙大子公司", "华东/浙大地区商家"),
        ("深圳、供应链、山东子公司", "深圳、供应链、山东地区商家"),
        ("中山子公司", "中山地区商家"),
        ("广州子公司", "广州地区商家"),
        ("佛山子公司", "佛山地区商家"),
        ("泉州子公司", "泉州地区商家"),
        ("福建子公司", "福建地区商家"),
        ("山东子公司", "山东地区商家"),
        ("浙大子公司", "浙大地区商家"),
        ("三家子公司", "三家地区商家"),
        ("可以通过的子公司", "可以通过的地区商家"),
        ("才能通过的子公司", "才能通过的地区商家"),
        ("的子公司有", "的地区商家有"),
        ("的子公司（", "的地区商家（"),
        ("子公司除外", "地区商家除外"),
        ("子公司口径", "地区商家口径"),
        ("子公司单证", "地区商家单证"),
        ("| 子公司 |", "| 适用地区商家 |"),
        ("广州&佛山&中山", "广州/佛山/中山地区商家"),
        ("广州佛山中山", "广州/佛山/中山地区商家"),
        ("子公司", "地区商家"),  # 兜底，避免残留
    ]
    for a, b in reps:
        text = text.replace(a, b)
    return text


def rewrite_region_cell(value: str) -> str:
    """Excel「子公司」列单元格 → 地区商家口径。"""
    v = clean_text(value)
    mapping = {
        "子公司": "适用地区商家",
        "佛山中山": "佛山/中山地区商家",
        "佛山": "佛山地区商家",
        "中山": "中山地区商家",
        "广州": "广州地区商家",
        "广州佛山中山": "广州/佛山/中山地区商家",
        "广州&佛山&中山": "广州/佛山/中山地区商家",
    }
    if v in mapping:
        return mapping[v]
    return rewrite_merchant_terms(v)


def is_cta(line: str) -> bool:
    return any(p.search(line) for p in CTA_PATTERNS)


def md_escape_cell(s: str) -> str:
    return clean_text(s).replace("|", "\\|").replace("\n", "<br>")


def write(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text.rstrip() + "\n", encoding="utf-8")


# ---------------------------------------------------------------------------
# Word helpers
# ---------------------------------------------------------------------------

def iter_block_items(parent: DocumentType):
    """Yield paragraphs and tables in document order."""
    body = parent.element.body
    for child in body.iterchildren():
        if child.tag == qn("w:p"):
            yield Paragraph(child, parent)
        elif child.tag == qn("w:tbl"):
            yield Table(child, parent)


def table_to_md(table: Table) -> str:
    rows = []
    for row in table.rows:
        cells = [md_escape_cell(c.text) for c in row.cells]
        rows.append(cells)
    if not rows:
        return ""
    cols = max(len(r) for r in rows)
    norm = [r + [""] * (cols - len(r)) for r in rows]
    lines = [
        "| " + " | ".join(norm[0]) + " |",
        "| " + " | ".join(["---"] * cols) + " |",
    ]
    for r in norm[1:]:
        lines.append("| " + " | ".join(r) + " |")
    return "\n".join(lines)


def convert_300_questions() -> list[Path]:
    doc = Document(SRC / "跨境外贸财税300问final.docx")
    # Parse into {qnum: {title, body_lines}}
    questions: dict[int, dict] = {}
    current: int | None = None
    disclaimer_lines: list[str] = []
    collecting_disclaimer = True

    for block in iter_block_items(doc):
        if isinstance(block, Table):
            if current is None:
                continue
            md = table_to_md(block)
            if md:
                questions[current]["body"].append(md)
                questions[current]["body"].append("")
            continue

        text = clean_text(block.text)
        if not text:
            continue

        style = block.style.name if block.style else "Normal"
        if style == "Heading 1" and re.match(r"^第\s*\d+\s*问", text):
            collecting_disclaimer = False
            m = re.match(r"^第\s*(\d+)\s*问\s*[:：]?\s*(.*)$", text)
            assert m
            qn_ = int(m.group(1))
            title = m.group(2).strip() or text
            current = qn_
            questions[qn_] = {"title": title, "raw": text, "body": []}
            continue

        if collecting_disclaimer and text not in ("目录",) and not text.startswith("第"):
            if text.startswith("第一部分") or text.startswith("第二部分"):
                collecting_disclaimer = False
                continue
            if text == "免责声明" or disclaimer_lines or text.startswith("财税问题"):
                disclaimer_lines.append(text)
            continue

        # skip repeated section banners between questions
        if re.match(r"^第[一二三四五]部分", text) or text in {
            "目录",
            "第1～80问",
            "第81～160问",
            "第161～210问",
            "第211～250问",
            "第251～300问",
        }:
            continue

        if current is None:
            continue
        if is_cta(text):
            continue
        questions[current]["body"].append(text)

    out_files: list[Path] = []
    index_lines = [
        "# 跨境外贸财税300问（索引）",
        "",
        f"> 更新：{AS_OF}",
        SCOPE,
        "> 原书免责声明摘要：解决方案随企业情形与法规变化而变化；内容仅供一般参考，不构成官方意见。",
        "",
        "## 分篇文件（请上传分篇，勿只传本索引）",
        "",
    ]

    for start, end, slug, label in PARTS:
        lines = [
            f"# 跨境外贸财税300问 · {label}",
            "",
            f"> 更新：{AS_OF} ｜来源：跨境外贸财税300问final.docx",
            SCOPE,
            f"> **范围**：仅本篇第{start}～{end}问。回答时勿串用其他篇章结论，除非用户明确跨主题。",
            "",
        ]
        for n in range(start, end + 1):
            q = questions.get(n)
            if not q:
                continue
            title = q["title"]
            lines.append(f"## 第{n}问：{title}")
            lines.append("")
            body = q["body"]
            # ensure answer starts clearly
            if body and not body[0].startswith("答"):
                lines.append("**答：**")
                lines.append("")
            for para in body:
                if para.startswith("|"):
                    lines.append(para)
                elif para.startswith("答：") or para.startswith("答:"):
                    lines.append(f"**{para[:2]}**{para[2:]}")
                else:
                    lines.append(para)
                lines.append("")
            lines.append("---")
            lines.append("")
        path = OUT / "300问" / f"{slug}.md"
        write(path, "\n".join(lines))
        out_files.append(path)
        index_lines.append(f"- [{label}](300问/{slug}.md)")

    index_lines += ["", "## 按题号速查", ""]
    for start, end, slug, label in PARTS:
        index_lines.append(f"- 第{start}～{end}问 → `{slug}.md`（{label.split('：')[-1]}）")
    write(OUT / "300问" / "00-索引.md", "\n".join(index_lines))
    out_files.append(OUT / "300问" / "00-索引.md")
    print(f"300问: {len(questions)} questions → {len(PARTS)} parts")
    return out_files


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def fill_down(rows: list[list], cols: list[int]) -> list[list]:
    last = {c: "" for c in cols}
    out = []
    for row in rows:
        r = list(row)
        for c in cols:
            if c < len(r) and r[c] not in (None, ""):
                last[c] = r[c]
            elif c < len(r):
                r[c] = last[c]
            else:
                r.extend([""] * (c + 1 - len(r)))
                r[c] = last[c]
        out.append(r)
    return out


def sheet_to_rows(ws) -> list[list[str]]:
    rows = []
    for row in ws.iter_rows(values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        rows.append(["" if v is None else clean_text(str(v)) for v in row])
    return rows


def rows_to_md_table(rows: list[list[str]]) -> str:
    if not rows:
        return ""
    cols = max(len(r) for r in rows)
    norm = [r + [""] * (cols - len(r)) for r in rows]
    # trim trailing empty columns
    while cols > 1 and all(not r[cols - 1] for r in norm):
        cols -= 1
        norm = [r[:cols] for r in norm]
    lines = [
        "| " + " | ".join(md_escape_cell(c) for c in norm[0]) + " |",
        "| " + " | ".join(["---"] * cols) + " |",
    ]
    for r in norm[1:]:
        lines.append("| " + " | ".join(md_escape_cell(c) for c in r) + " |")
    return "\n".join(lines)


def convert_audit_rules() -> Path:
    wb = openpyxl.load_workbook(SRC / "审核规则-佛山中山.xlsx", data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = sheet_to_rows(ws)
    # fill 单证环节 / 审核点 / 地区
    rows = fill_down(rows, [0, 1, 2])
    # 表头与地区列改写；正文去掉「子公司」字眼
    if rows:
        rows[0] = [rewrite_region_cell(c) if i == 2 or c == "子公司" else c for i, c in enumerate(rows[0])]
        if len(rows[0]) > 2:
            rows[0][2] = "适用地区商家"
        for i in range(1, len(rows)):
            r = rows[i] + [""] * 5
            r[2] = rewrite_region_cell(r[2])
            r[3] = rewrite_merchant_terms(r[3])
            r[4] = rewrite_merchant_terms(r[4])
            rows[i] = r[:5]
    lines = [
        "# 出口退税单证审核规则（佛山 / 中山地区商家）",
        "",
        f"> 更新：{AS_OF} ｜来源：审核规则-佛山中山.xlsx",
        SCOPE,
        "> **范围**：仅**佛山、中山地区商家**单证审核口径。"
        "与其他地区商家要求不同，回答时请先确认商家所属地区。",
        "",
        "## 总表",
        "",
        rows_to_md_table(rows),
        "",
        "## 分条细则（便于检索）",
        "",
    ]
    groups: dict[tuple[str, str], list[tuple[str, str, str]]] = defaultdict(list)
    for r in rows[1:]:
        link, point, sub, rule, special = (r + [""] * 5)[:5]
        groups[(link, point)].append((sub, rule, special))

    for (link, point), items in groups.items():
        title = f"{link}" if point in ("/", "", "／") else f"{link} · {point}"
        lines.append(f"### {title}")
        lines.append("")
        for sub, rule, special in items:
            lines.append(f"- **适用地区商家**：{sub}")
            lines.append(f"  - **审核规则**：{rule}")
            if special:
                lines.append(f"  - **特殊要求**：{special}")
        lines.append("")

    path = OUT / "审核规则" / "佛山中山-单证审核规则.md"
    write(path, rewrite_merchant_terms("\n".join(lines)))
    print("audit rules →", path.name)
    return path


def convert_first_declare_list() -> Path:
    wb = openpyxl.load_workbook(SRC / "首次申报出口退税资料清单5.0.xlsx", data_only=True)
    # Prefer sheet with 填报说明
    sheet_name = "一般调评" if "一般调评" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = fill_down(sheet_to_rows(ws), [0])
    lines = [
        "# 首次申报出口退税资料清单（一般调评）",
        "",
        f"> 更新：{AS_OF} ｜来源：首次申报出口退税资料清单5.0.xlsx（{sheet_name}）",
        SCOPE,
        "> **范围**：出口企业首次申报退税调查评估/实地核查常见资料要求。地区税局口径可能更细，以当地要求为准。",
        "",
        "## 资料总表",
        "",
        rows_to_md_table(rows),
        "",
        "## 按项目分列",
        "",
    ]
    current = None
    for r in rows[1:]:
        proj, name, note = (r + [""] * 3)[:3]
        if proj != current:
            current = proj
            lines.append(f"### {proj}")
            lines.append("")
        lines.append(f"- **{name}**")
        if note:
            lines.append(f"  - 填报说明：{note}")
    lines.append("")

    # optional external short list
    if "一般调评 对外" in wb.sheetnames:
        rows2 = fill_down(sheet_to_rows(wb["一般调评 对外"]), [0])
        lines += [
            "## 附录：对外简版清单（仅项目与资料名称）",
            "",
            rows_to_md_table(rows2),
            "",
        ]

    path = OUT / "首次申报" / "首次申报出口退税资料清单.md"
    write(path, "\n".join(lines))
    print("first declare →", path.name)
    return path


# ---------------------------------------------------------------------------
# XMind
# ---------------------------------------------------------------------------

XMIND_NS = "{urn:xmind:xmap:xmlns:content:2.0}"


def xmind_title(topic) -> str:
    el = topic.find(f"{XMIND_NS}title")
    if el is None or not el.text:
        return ""
    return rewrite_merchant_terms(clean_text(el.text.replace("\r", "\n")))


def xmind_children(topic):
    children = topic.find(f"{XMIND_NS}children")
    if children is None:
        return []
    out = []
    for topics in children.findall(f"{XMIND_NS}topics"):
        out.extend(topics.findall(f"{XMIND_NS}topic"))
    return out


def topic_to_md(topic, depth: int, lines: list[str]) -> None:
    title = xmind_title(topic)
    kids = xmind_children(topic)
    if not title and not kids:
        return

    # Empty structural nodes: promote children (avoid “(未命名节点)”)
    if not title:
        for child in kids:
            topic_to_md(child, depth, lines)
        return

    # Multi-line titles: first line as heading/bullet, rest as body
    parts = [p.strip() for p in title.split("\n") if p.strip()]
    head = parts[0]
    rest = parts[1:]

    if depth <= 2:
        level = min(depth + 1, 4)  # sheet root already H1; L1->H2, L2->H3
        lines.append("#" * level + f" {head}")
        lines.append("")
        for p in rest:
            lines.append(p)
            lines.append("")
    else:
        indent = "  " * (depth - 3)
        lines.append(f"{indent}- {head}")
        for p in rest:
            lines.append(f"{indent}  - {p}")

    for child in kids:
        topic_to_md(child, depth + 1, lines)

    if depth <= 2 and kids:
        lines.append("")


def convert_xmind() -> list[Path]:
    path = SRC / "备案单证规则-20240806.xmind"
    with zipfile.ZipFile(path) as z:
        root = ET.fromstring(z.read("content.xml"))

    out_files = []
    index = [
        "# 备案单证审核规则（脑图清洗）",
        "",
        f"> 更新：{AS_OF} ｜来源：备案单证规则-20240806.xmind",
        SCOPE,
        "> **范围**：出口退税备案单证审核细则。按单证类型分篇；"
        "回答时锁定单证类型，并区分**佛山/中山等地区商家**与**其他地区商家**要求。",
        "",
        "## 分篇",
        "",
    ]

    sheets = root.findall(f".//{XMIND_NS}sheet")
    for i, sheet in enumerate(sheets, 1):
        st_el = sheet.find(f"{XMIND_NS}title")
        sheet_title = clean_text(st_el.text) if st_el is not None and st_el.text else f"sheet{i}"
        root_topic = sheet.find(f"{XMIND_NS}topic")
        # Prefer sheet title as doc title; root topic title often equals sheet
        doc_title = sheet_title
        lines = [
            f"# 备案单证规则 · {doc_title}",
            "",
            f"> 更新：{AS_OF} ｜来源：备案单证规则-20240806.xmind / {sheet_title}",
            SCOPE,
            f"> **范围**：仅「{doc_title}」相关审核点。"
            "勿把海运规则答到空运/快递；并区分佛山/中山地区商家与其他地区商家要求。",
            "",
        ]
        if root_topic is not None:
            # skip duplicating root title if same as sheet
            rt = xmind_title(root_topic)
            children = xmind_children(root_topic)
            if rt and rt != sheet_title and rt not in sheet_title:
                lines.append(f"## 总览：{rt.split(chr(10))[0]}")
                lines.append("")
            for child in children:
                topic_to_md(child, depth=1, lines=lines)

        slug = f"{i:02d}-{sheet_title}"
        slug = re.sub(r"[\\/:*?\"<>|]+", "-", slug)
        out = OUT / "备案单证规则" / f"{slug}.md"
        write(out, rewrite_merchant_terms("\n".join(lines)))
        out_files.append(out)
        index.append(f"- [{sheet_title}](备案单证规则/{out.name})")

    write(OUT / "备案单证规则" / "00-索引.md", rewrite_merchant_terms("\n".join(index) + "\n"))
    out_files.append(OUT / "备案单证规则" / "00-索引.md")
    print(f"xmind: {len(sheets)} sheets")
    return out_files


def write_readme(paths: list[Path]) -> None:
    rel = [p.relative_to(OUT).as_posix() for p in paths if p.name != "README.md"]
    text = f"""# 道一跨境电商知识库（Markdown 清洗版）

> 更新：{AS_OF}
> 源文件夹：桌面 `DIFY知识库/道一跨境电商知识库`

## 上传到 Dify 的建议

1. **上传下列分篇**，不要把所有内容糊成一个超大 MD。
2. 切片优先按 `##`（300问每问一节；审核规则按审核点；备案单证按二级标题）。
3. **不要**再上传源 Word/Excel/XMind（已清洗）；也不要上传本目录外的「平台税政策汇总」大合集到同一库（易串台）。
4. 原 Word 中的示意图未嵌入；文字结论已保留。若某问依赖截图，请对照原 Word。

## 文件清单

### 跨境外贸财税300问
- `300问/01-财税合规篇.md` … `05-外汇关务篇.md`（共5篇，300问）
- `300问/00-索引.md`（可选，可不上传）

### 审核与资料清单
- `审核规则/佛山中山-单证审核规则.md`
- `首次申报/首次申报出口退税资料清单.md`

### 备案单证规则（原 XMind）
- `备案单证规则/01-合同.md` … 等按运单类型分篇
- `备案单证规则/00-索引.md`（可选）

## 生成文件数

共 {len(rel)} 个 md（含索引）。
"""
    write(OUT / "README.md", text)


def sync_desktop() -> None:
    if DESKTOP_MD.exists():
        shutil.rmtree(DESKTOP_MD)
    shutil.copytree(OUT, DESKTOP_MD)
    print("synced →", DESKTOP_MD)


def main():
    assert SRC.exists(), SRC
    if OUT.exists():
        shutil.rmtree(OUT)
    OUT.mkdir(parents=True)

    paths: list[Path] = []
    paths += convert_300_questions()
    paths.append(convert_audit_rules())
    paths.append(convert_first_declare_list())
    paths += convert_xmind()
    write_readme(paths)
    sync_desktop()

    # size summary
    mds = sorted(OUT.rglob("*.md"))
    print("total md", len(mds))
    for p in mds:
        print(f"  {p.relative_to(OUT)} ({p.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
